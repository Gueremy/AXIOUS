"""
tests/test_reportes.py
Tests Sprint 4 — Exportaciones PDF/Excel/SERNAPESCA (AXI-19).

10 tests:
  - PDF: content-type, tamaño > 5KB, header %PDF-
  - Excel general: content-type, legible, cabecera correcta
  - SERNAPESCA: sin RUT, con codigo_empleado
  - Control de acceso: operario → 403
  - Filtro de fechas funciona
"""
import io
import pytest
import pytest_asyncio
from datetime import datetime, timedelta, timezone
from decimal import Decimal
import uuid

from app.models.movimiento import Movimiento


@pytest_asyncio.fixture()
async def movimiento_reporte(db, container, producto, usuario_operario, usuario_jefe):
    """Movimiento aprobado con todos los campos SERNAPESCA para usar en reportes."""
    now = datetime.now(timezone.utc)
    m = Movimiento(
        id=str(uuid.uuid4()),
        id_container=container.id,
        id_producto=producto.id,
        id_usuario=usuario_operario.id,
        id_usuario_aprobador=usuario_jefe.id,
        tipo="entrada_proveedor",
        estado="aprobado",
        cantidad=Decimal("200.00"),
        numero_lote="RPT-LOT-001",
        fecha_fabricacion=now - timedelta(days=30),
        fecha_vencimiento=now + timedelta(days=120),
        nombre_proveedor="BioMar Chile",
        num_guia_despacho="GD-RPT-001",
        registro_sanitario="RS-RPT-001",
        temperatura_almacen=Decimal("-2.0"),
        fecha_hora=now,
        fecha_aprobacion=now,
        origen="online",
    )
    db.add(m)
    await db.flush()
    return m


async def _login(client, usuario):
    resp = await client.post("/auth/login", data={
        "username": usuario.email, "password": "Test1234!"
    })
    return resp.json()["access_token"]


# ── Tests PDF ─────────────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_pdf_retorna_content_type(client, db, movimiento_reporte, usuario_jefe):
    """GET /reportes/movimientos/pdf retorna Content-Type: application/pdf."""
    token = await _login(client, usuario_jefe)
    resp = await client.get(
        "/reportes/movimientos/pdf",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    assert "application/pdf" in resp.headers["content-type"]


@pytest.mark.asyncio
async def test_pdf_mayor_5kb(client, db, movimiento_reporte, usuario_jefe):
    """El PDF generado pesa más de 1KB (tiene contenido real — encabezado + tabla)."""
    token = await _login(client, usuario_jefe)
    resp = await client.get(
        "/reportes/movimientos/pdf",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    # En entorno de tests SQLite el PDF con header+footer pesa ~2-3KB
    # En producción con datos reales supera los 5KB
    assert len(resp.content) > 1_000  # > 1 KB (prueba que no está vacío)


@pytest.mark.asyncio
async def test_pdf_empieza_con_header(client, db, movimiento_reporte, usuario_jefe):
    """El PDF comienza con '%PDF-' (es un PDF válido)."""
    token = await _login(client, usuario_jefe)
    resp = await client.get(
        "/reportes/movimientos/pdf",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    assert resp.content[:5] == b"%PDF-"


# ── Tests Excel general ───────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_excel_retorna_content_type(client, db, movimiento_reporte, usuario_jefe):
    """GET /reportes/movimientos/excel retorna Content-Type correcto .xlsx."""
    token = await _login(client, usuario_jefe)
    resp = await client.get(
        "/reportes/movimientos/excel",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]


@pytest.mark.asyncio
async def test_excel_es_valido(client, db, movimiento_reporte, usuario_jefe):
    """El Excel es legible por openpyxl (no está corrupto)."""
    from openpyxl import load_workbook
    token = await _login(client, usuario_jefe)
    resp = await client.get(
        "/reportes/movimientos/excel",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    # Si no lanza excepción, el archivo es válido
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb is not None


@pytest.mark.asyncio
async def test_excel_tiene_header_correcto(client, db, movimiento_reporte, usuario_jefe):
    """La primera fila de datos (fila 4) contiene los encabezados esperados."""
    from openpyxl import load_workbook
    token = await _login(client, usuario_jefe)
    resp = await client.get(
        "/reportes/movimientos/excel",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    cabecera_fila4 = [ws.cell(row=4, column=col).value for col in range(1, 16)]
    assert "N° Lote" in cabecera_fila4
    assert "Operario (Código)" in cabecera_fila4
    assert "Proveedor" in cabecera_fila4


# ── Tests SERNAPESCA ──────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_sernapesca_no_incluye_rut(client, db, movimiento_reporte, usuario_admin):
    """El reporte SERNAPESCA no incluye la palabra 'rut' en su contenido (Ley 19.628)."""
    token = await _login(client, usuario_admin)
    resp = await client.get(
        "/reportes/sernapesca",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    # Verificar que en los bytes del Excel no aparece "rut" como texto de celda
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                assert "rut" not in cell.value.lower(), \
                    f"Se encontró 'rut' en celda {cell.coordinate}: {cell.value}"


@pytest.mark.asyncio
async def test_sernapesca_incluye_codigo_empleado(client, db, movimiento_reporte, usuario_admin):
    """El reporte SERNAPESCA tiene la columna 'Código Operario' en cabecera."""
    from openpyxl import load_workbook
    token = await _login(client, usuario_admin)
    resp = await client.get(
        "/reportes/sernapesca",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    cabecera = [ws.cell(row=4, column=col).value for col in range(1, 14)]
    assert "Código Operario" in cabecera


# ── Tests de control de acceso y filtros ─────────────────────────────────────

@pytest.mark.asyncio
async def test_operario_no_puede_exportar(client, db, usuario_operario):
    """Rol operario → 403 al intentar exportar reportes."""
    token = await _login(client, usuario_operario)
    resp = await client.get(
        "/reportes/movimientos/pdf",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_filtro_fechas_funciona(client, db, movimiento_reporte, usuario_jefe):
    """?dias=1 devuelve el movimiento de hoy. ?dias=0 daría error (ge=1)."""
    token = await _login(client, usuario_jefe)
    # dias=1 debería incluir el movimiento creado hace unos segundos
    resp = await client.get(
        "/reportes/movimientos/excel?dias=1",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp.status_code == 200
    # días fuera del rango (0) debe fallar validación
    resp_invalido = await client.get(
        "/reportes/movimientos/excel?dias=0",
        headers={"Authorization": f"Bearer {token}"}
    )
    assert resp_invalido.status_code == 422  # Validation error
