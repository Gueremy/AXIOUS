"""
app/services/reportes.py
Generación de reportes PDF y Excel — Sprint 4 (AXI-19).

Funciones:
  generar_pdf_movimientos()      → bytes PDF (reportlab)
  generar_excel_movimientos()    → bytes Excel .xlsx (openpyxl)
  generar_reporte_sernapesca()   → bytes Excel .xlsx formato SERNAPESCA

⚠️ NUNCA incluir RUT en ningún reporte (Ley 19.628)
   Usar siempre codigo_empleado
"""
import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ── Colores corporativos Skretting ───────────────────────────────────────────
AZUL_SKRETTING = colors.HexColor("#1A3A6B")
VERDE_SKRETTING = colors.HexColor("#0E7A3D")
GRIS_CABECERA = colors.HexColor("#F0F4F8")
ROJO_ALERTA = colors.HexColor("#D32F2F")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _fmt_fecha(dt) -> str:
    """Formatea datetime a 'DD/MM/YYYY HH:MM'. None → '—'."""
    if not dt:
        return "—"
    if hasattr(dt, "strftime"):
        return dt.strftime("%d/%m/%Y %H:%M")
    return str(dt)


def _fmt_fecha_solo(dt) -> str:
    """Formatea datetime a 'DD/MM/YYYY'. None → '—'."""
    if not dt:
        return "—"
    if hasattr(dt, "strftime"):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def _estilo_excel_cabecera(ws, fila: int, num_cols: int, color_hex: str = "1A3A6B"):
    """Aplica estilo de cabecera azul a la fila indicada."""
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    alin = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = Border(
        bottom=Side(style="thin", color="FFFFFF"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alin
        cell.border = borde


def _autoajuste_columnas(ws):
    """Ajusta el ancho de cada columna al contenido máximo."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ── PDF ──────────────────────────────────────────────────────────────────────

def generar_pdf_movimientos(
    movimientos: list[dict],
    nombre_sede: str = "Sede Skretting",
    desde: Optional[datetime] = None,
    hasta: Optional[datetime] = None,
) -> bytes:
    """
    Genera un PDF con tabla de movimientos SERNAPESCA.
    Retorna los bytes del PDF.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    style_titulo = ParagraphStyle(
        "titulo",
        parent=styles["Title"],
        textColor=AZUL_SKRETTING,
        fontSize=16,
        spaceAfter=6,
    )
    style_subtitulo = ParagraphStyle(
        "subtitulo",
        parent=styles["Normal"],
        textColor=colors.grey,
        fontSize=10,
        spaceAfter=4,
    )
    style_celda = ParagraphStyle(
        "celda",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
    )

    # ── Encabezado ────────────────────────────────────────────────────────────
    story = []
    story.append(Paragraph("Skretting Chile Ltda.", style_titulo))
    story.append(Paragraph(f"Reporte de Movimientos — {nombre_sede}", style_subtitulo))

    rango = ""
    if desde and hasta:
        rango = f"Período: {_fmt_fecha_solo(desde)} — {_fmt_fecha_solo(hasta)}"
    elif desde:
        rango = f"Desde: {_fmt_fecha_solo(desde)}"
    rango += f"  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    story.append(Paragraph(rango, style_subtitulo))
    story.append(HRFlowable(width="100%", thickness=2, color=AZUL_SKRETTING))
    story.append(Spacer(1, 0.4 * cm))

    # ── Tabla ────────────────────────────────────────────────────────────────
    headers = [
        "Fecha/Hora", "Tipo", "Estado", "Lote", "Cantidad",
        "Proveedor", "N° Guía", "Reg. Sanitario",
        "F. Fabricación", "F. Vencimiento", "Temp. (°C)",
        "Container", "Galpón", "Operario",
    ]

    data = [headers]
    for m in movimientos:
        data.append([
            Paragraph(_fmt_fecha(m.get("fecha_hora")), style_celda),
            Paragraph(str(m.get("tipo", "—")), style_celda),
            Paragraph(str(m.get("estado", "—")), style_celda),
            Paragraph(str(m.get("numero_lote", "—")), style_celda),
            Paragraph(f"{m.get('cantidad', 0)} {m.get('unidad_medida', '')}", style_celda),
            Paragraph(str(m.get("nombre_proveedor") or "—"), style_celda),
            Paragraph(str(m.get("num_guia_despacho") or "—"), style_celda),
            Paragraph(str(m.get("registro_sanitario") or "—"), style_celda),
            Paragraph(_fmt_fecha_solo(m.get("fecha_fabricacion")), style_celda),
            Paragraph(_fmt_fecha_solo(m.get("fecha_vencimiento")), style_celda),
            Paragraph(str(m.get("temperatura_almacen") or "—"), style_celda),
            Paragraph(str(m.get("container_codigo") or "—"), style_celda),
            Paragraph(str(m.get("galpon_nombre") or "—"), style_celda),
            Paragraph(str(m.get("codigo_empleado") or "—"), style_celda),
        ])

    col_widths = [
        3.0 * cm,  # fecha
        2.5 * cm,  # tipo
        2.0 * cm,  # estado
        2.5 * cm,  # lote
        2.0 * cm,  # cantidad
        3.5 * cm,  # proveedor
        2.5 * cm,  # guía
        2.5 * cm,  # reg sanitario
        2.5 * cm,  # f fabricacion
        2.5 * cm,  # f vencimiento
        2.0 * cm,  # temperatura
        2.5 * cm,  # container
        2.5 * cm,  # galpon
        2.5 * cm,  # operario
    ]

    tabla = Table(data, colWidths=col_widths, repeatRows=1)
    tabla.setStyle(TableStyle([
        # Cabecera
        ("BACKGROUND",   (0, 0), (-1, 0), AZUL_SKRETTING),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 7),
        ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        # Filas pares
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CABECERA]),
        # Bordes
        ("GRID",         (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("LINEBELOW",    (0, 0), (-1, 0), 1.5, AZUL_SKRETTING),
        # Padding
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))

    story.append(tabla)
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        f"Total registros: {len(movimientos)} — Cumplimiento Ley 19.628: RUT no incluido",
        style_subtitulo
    ))

    doc.build(story)
    return buffer.getvalue()


# ── Excel General ────────────────────────────────────────────────────────────

def generar_excel_movimientos(
    movimientos: list[dict],
    nombre_sede: str = "Sede Skretting",
) -> bytes:
    """
    Genera un Excel .xlsx con tabla de movimientos auditable.
    Retorna los bytes del archivo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    ws.sheet_view.showGridLines = True

    # Metadata del archivo
    wb.properties.title = "Reporte de Movimientos — Skretting Chile"
    wb.properties.creator = "Sistema AXIOUS — Skretting"

    # Fila 1: Título
    ws.merge_cells("A1:N1")
    ws["A1"] = f"Reporte de Movimientos — {nombre_sede}"
    ws["A1"].font = Font(bold=True, size=14, color="1A3A6B")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Fila 2: Generado
    ws.merge_cells("A2:N2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Ley 19.628: RUT no incluido"
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])  # Fila 3 vacía

    # Fila 4: Cabecera
    CABECERA = [
        "Fecha/Hora", "Tipo", "Estado", "N° Lote", "Cantidad", "Unidad",
        "Proveedor", "N° Guía Despacho", "Registro Sanitario",
        "F. Fabricación", "F. Vencimiento", "Temp. Almacén (°C)",
        "Container", "Galpón", "Operario (Código)",
    ]
    ws.append(CABECERA)
    _estilo_excel_cabecera(ws, 4, len(CABECERA))
    ws.row_dimensions[4].height = 30

    # Filas de datos
    fill_par = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    for i, m in enumerate(movimientos):
        fila = [
            _fmt_fecha(m.get("fecha_hora")),
            m.get("tipo", "—"),
            m.get("estado", "—"),
            m.get("numero_lote", "—"),
            float(m.get("cantidad", 0)),
            m.get("unidad_medida", "—"),
            m.get("nombre_proveedor") or "—",
            m.get("num_guia_despacho") or "—",
            m.get("registro_sanitario") or "—",
            _fmt_fecha_solo(m.get("fecha_fabricacion")),
            _fmt_fecha_solo(m.get("fecha_vencimiento")),
            float(m.get("temperatura_almacen")) if m.get("temperatura_almacen") is not None else "—",
            m.get("container_codigo") or "—",
            m.get("galpon_nombre") or "—",
            m.get("codigo_empleado") or "—",
        ]
        ws.append(fila)
        # Alternar color de fila
        if i % 2 == 1:
            for col in range(1, len(CABECERA) + 1):
                ws.cell(row=5 + i, column=col).fill = fill_par

    # Autoajuste y freeze header
    _autoajuste_columnas(ws)
    ws.freeze_panes = "A5"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Excel SERNAPESCA ─────────────────────────────────────────────────────────

def generar_reporte_sernapesca(
    movimientos: list[dict],
    nombre_sede: str = "Sede Skretting",
) -> bytes:
    """
    Genera Excel con el formato específico exigido por SERNAPESCA.
    Columnas: las definidas por la normativa.
    ⚠️ NUNCA incluir RUT (Ley 19.628) — usar codigo_empleado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SERNAPESCA"

    wb.properties.title = "Reporte SERNAPESCA — Skretting Chile"

    # Encabezado institucional
    ws.merge_cells("A1:L1")
    ws["A1"] = "REGISTRO DE MOVIMIENTO DE ALIMENTOS — SERNAPESCA"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="0E7A3D", end_color="0E7A3D", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Empresa: Skretting Chile Ltda. | Sede: {nombre_sede} | Generado: {datetime.now().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])

    # Columnas SERNAPESCA (normativa)
    CABECERA_SRN = [
        "Fecha Movimiento",
        "Tipo Movimiento",
        "N° Lote",
        "Cantidad",
        "Unidad Medida",
        "Nombre Proveedor",
        "N° Guía Despacho",
        "Registro Sanitario",
        "Fecha Fabricación",
        "Fecha Vencimiento",
        "Temperatura Almacén (°C)",
        "Código Operario",    # NUNCA RUT — Ley 19.628
    ]
    ws.append(CABECERA_SRN)
    _estilo_excel_cabecera(ws, 4, len(CABECERA_SRN), color_hex="0E7A3D")
    ws.row_dimensions[4].height = 32

    fill_par = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    for i, m in enumerate(movimientos):
        fila = [
            _fmt_fecha(m.get("fecha_hora")),
            m.get("tipo", "—"),
            m.get("numero_lote", "—"),
            float(m.get("cantidad", 0)),
            m.get("unidad_medida", "—"),
            m.get("nombre_proveedor") or "—",
            m.get("num_guia_despacho") or "—",
            m.get("registro_sanitario") or "—",
            _fmt_fecha_solo(m.get("fecha_fabricacion")),
            _fmt_fecha_solo(m.get("fecha_vencimiento")),
            float(m.get("temperatura_almacen")) if m.get("temperatura_almacen") is not None else "—",
            m.get("codigo_empleado") or "—",   # NUNCA rut
        ]
        ws.append(fila)
        if i % 2 == 1:
            for col in range(1, len(CABECERA_SRN) + 1):
                ws.cell(row=5 + i, column=col).fill = fill_par

    _autoajuste_columnas(ws)
    ws.freeze_panes = "A5"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
